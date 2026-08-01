"""
Animal master-data importer — the first of the "connectors" (per the user's
plan: Excel first, other systems/formats later). Deliberately tolerant of
whatever a real farm's existing spreadsheet actually looks like: column
names are matched against a set of common aliases (case-insensitive, blank/
underscore/space-insensitive) rather than requiring an exact template match.

Both `preview()` and `commit()` share this module's row-parsing logic so the
preview a farm sees is guaranteed to match what actually happens on commit —
the only difference is preview never calls .save().
"""
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl

from apps.animals.models import Animal, AnimalStatus, Sex
from apps.tenants.models import Breed

# Each canonical field maps to the column-header variants we'll recognize.
# Matching is case-insensitive and ignores spaces/underscores/hyphens, so
# "Tag No", "tag_no", "TAG-NO" all resolve to the same alias entry below.
FIELD_ALIASES = {
    "tag_number": ["tag number", "tag no", "tag", "ear tag", "animal id", "animal tag", "id"],
    "name": ["name", "animal name"],
    "breed": ["breed"],
    "sex": ["sex", "gender"],
    "date_of_birth": ["date of birth", "dob", "birth date", "birthdate"],
    "status": ["status"],
    "weight_kg": ["weight", "weight kg", "weight (kg)"],
    "lactation_number": ["lactation number", "lactation no", "lactation"],
    "dam_tag": ["dam", "dam tag", "mother", "dam tag number"],
    "sire_tag": ["sire", "sire tag", "father", "bull tag", "bull id"],
    "purchase_date": ["purchase date"],
    "purchase_price": ["purchase price", "price"],
    "notes": ["notes", "remarks", "comments"],
}

REQUIRED_FIELDS = ["tag_number"]

VALID_STATUSES = {c[0] for c in AnimalStatus.choices}
VALID_SEXES = {c[0] for c in Sex.choices}


def _normalize_header(h):
    if h is None:
        return ""
    return re.sub(r"[\s_\-]+", " ", str(h).strip().lower()).strip()


def build_column_map(header_row):
    """Maps canonical field name -> the actual column index found in this file's header row."""
    normalized = {_normalize_header(h): i for i, h in enumerate(header_row)}
    alias_lookup = {}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            alias_lookup[_normalize_header(alias)] = field

    column_map = {}
    for norm_header, idx in normalized.items():
        field = alias_lookup.get(norm_header)
        if field and field not in column_map:
            column_map[field] = idx
    return column_map


def _parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    raise ValueError(f"unrecognized date format: {value!r}")


def _parse_decimal(value, field_name):
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        raise ValueError(f"{field_name} must be a number, got {value!r}")


def parse_workbook(file):
    """Returns (column_map, data_rows) or raises ValueError with a user-facing message."""
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        raise ValueError(f"Couldn't read this as an Excel file: {e}")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The file is empty.")

    header_row = rows[0]
    column_map = build_column_map(header_row)

    missing = [f for f in REQUIRED_FIELDS if f not in column_map]
    if missing:
        raise ValueError(
            f"Couldn't find a column for: {', '.join(missing)}. "
            f"Recognized header variants include: {', '.join(FIELD_ALIASES['tag_number'])}."
        )

    data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
    return column_map, data_rows


def process_rows(tenant, column_map, data_rows, user=None, commit=False):
    """
    Walks every data row, validates it, and (if commit=True) actually saves
    it. Always returns the same result shape whether previewing or
    committing, so the frontend preview table and the final outcome use
    identical rendering logic.

    Dam/sire matching checks animals already in the DB *and* animals created
    earlier in this same batch (so a spreadsheet listing a dam before her
    calves, in any order, still links correctly within one import) — but
    NOT animals appearing *later* in the batch (a dam must already exist or
    have been processed to be linkable; this is a one-pass importer, not a
    two-pass dependency resolver, which is a reasonable limit for how real
    spreadsheets are usually ordered — oldest/parent animals first).
    """
    results = []
    tag_to_animal = {}  # tag_number -> Animal instance, seeded with existing + filled in as we create

    existing_tags = set(
        Animal.objects.filter(tenant=tenant).values_list("tag_number", flat=True)
    )
    seen_tags_in_batch = set()

    def get_col(row, field):
        idx = column_map.get(field)
        if idx is None or idx >= len(row):
            return None
        val = row[idx]
        if isinstance(val, str):
            val = val.strip()
        return val if val != "" else None

    for i, row in enumerate(data_rows, start=2):  # row 1 is the header
        row_errors = []
        row_warnings = []

        tag_number = get_col(row, "tag_number")
        if not tag_number:
            results.append({"row": i, "status": "error", "errors": ["tag_number is required"], "data": None})
            continue
        tag_number = str(tag_number).strip()

        if tag_number in existing_tags:
            results.append({
                "row": i, "status": "error",
                "errors": [f"An animal with tag '{tag_number}' already exists in this farm — skipped."],
                "data": None,
            })
            continue
        if tag_number in seen_tags_in_batch:
            results.append({
                "row": i, "status": "error",
                "errors": [f"Tag '{tag_number}' appears more than once in this file — skipped."],
                "data": None,
            })
            continue

        sex_raw = get_col(row, "sex")
        sex = None
        if sex_raw:
            sex_norm = str(sex_raw).strip().lower()
            sex_alias = {"f": "female", "female": "female", "m": "male", "male": "male"}
            sex = sex_alias.get(sex_norm)
            if sex not in VALID_SEXES:
                row_errors.append(f"sex '{sex_raw}' not recognized (expected male/female)")
                sex = None
        sex = sex or Sex.FEMALE

        status_raw = get_col(row, "status")
        status = AnimalStatus.HEIFER
        if status_raw:
            status_norm = str(status_raw).strip().lower()
            if status_norm in VALID_STATUSES:
                status = status_norm
            else:
                row_errors.append(
                    f"status '{status_raw}' not recognized (expected one of: {', '.join(sorted(VALID_STATUSES))})"
                )

        dob = None
        try:
            dob = _parse_date(get_col(row, "date_of_birth"))
        except ValueError as e:
            row_errors.append(str(e))

        purchase_date = None
        try:
            purchase_date = _parse_date(get_col(row, "purchase_date"))
        except ValueError as e:
            row_errors.append(str(e))

        weight_kg = purchase_price = None
        try:
            weight_kg = _parse_decimal(get_col(row, "weight_kg"), "weight")
        except ValueError as e:
            row_errors.append(str(e))
        try:
            purchase_price = _parse_decimal(get_col(row, "purchase_price"), "purchase price")
        except ValueError as e:
            row_errors.append(str(e))

        lactation_raw = get_col(row, "lactation_number")
        lactation_number = 0
        if lactation_raw is not None:
            try:
                lactation_number = int(lactation_raw)
            except (ValueError, TypeError):
                row_errors.append(f"lactation number must be a whole number, got {lactation_raw!r}")

        breed_name = get_col(row, "breed")
        breed = None
        if breed_name:
            breed_name = str(breed_name).strip()
            breed = (
                Breed.objects.filter(name__iexact=breed_name, tenant=tenant).first()
                or Breed.objects.filter(name__iexact=breed_name, is_global=True).first()
            )
            # row_errors is already final by this point (nothing below adds an
            # error, only warnings) so this check keeps a row that ultimately
            # fails from leaving behind an orphan Breed it never ends up using.
            if not breed and commit and not row_errors:
                breed = Breed.objects.create(name=breed_name, tenant=tenant, species="cattle")
            elif not breed:
                row_warnings.append(f"breed '{breed_name}' not found — will be created on import")

        # A dam/sire tag "resolves" if it's either already in the DB or was
        # validated successfully earlier in this same batch. That second
        # check must work identically in preview and commit — otherwise
        # preview would show "not found" for a dam-before-calf ordering that
        # actually links up fine on commit (only tag_to_animal, the real
        # Animal-instance cache, is preview-mode-unavailable; the tag
        # existence itself is known regardless of commit mode).
        def tag_resolves(tag):
            return tag in seen_tags_in_batch or Animal.objects.filter(tenant=tenant, tag_number=tag).exists()

        dam_tag = get_col(row, "dam_tag")
        dam = None
        dam_resolved_tag = None
        if dam_tag:
            dam_tag = str(dam_tag).strip()
            if tag_resolves(dam_tag):
                dam_resolved_tag = dam_tag
                if commit:
                    dam = tag_to_animal.get(dam_tag) or Animal.objects.filter(tenant=tenant, tag_number=dam_tag).first()
            else:
                row_warnings.append(f"dam with tag '{dam_tag}' not found — left unlinked")

        sire_tag = get_col(row, "sire_tag")
        sire = None
        sire_resolved_tag = None
        sire_tag_text = ""
        if sire_tag:
            sire_tag = str(sire_tag).strip()
            if tag_resolves(sire_tag):
                sire_resolved_tag = sire_tag
                if commit:
                    sire = tag_to_animal.get(sire_tag) or Animal.objects.filter(tenant=tenant, tag_number=sire_tag).first()
            else:
                sire_tag_text = sire_tag  # preserved as free text, matching the model's existing fallback

        name = get_col(row, "name") or ""
        notes = get_col(row, "notes") or ""

        row_data = {
            "tag_number": tag_number, "name": name, "sex": sex, "status": status,
            "date_of_birth": dob, "weight_kg": weight_kg, "lactation_number": lactation_number,
            "breed": breed.name if breed else (breed_name or None),
            "dam_tag": dam_resolved_tag,
            "sire_tag": sire_resolved_tag or (sire_tag_text or None),
            "purchase_date": purchase_date, "purchase_price": purchase_price, "notes": notes,
        }

        if row_errors:
            results.append({"row": i, "status": "error", "errors": row_errors, "data": row_data})
            continue

        seen_tags_in_batch.add(tag_number)

        if commit:
            animal = Animal.objects.create(
                tenant=tenant, tag_number=tag_number, name=name, sex=sex, status=status,
                date_of_birth=dob, weight_kg=weight_kg, lactation_number=lactation_number,
                breed=breed, dam=dam, sire=sire, sire_tag=sire_tag_text,
                purchase_date=purchase_date, purchase_price=purchase_price, notes=notes,
            )
            tag_to_animal[tag_number] = animal

        results.append({
            "row": i, "status": "ok" if not row_warnings else "warning",
            "errors": [], "warnings": row_warnings, "data": row_data,
        })

    return results


def build_template():
    """A downloadable starting-point template — not required (aliases handle real-world sheets), just a convenience."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Animals"
    headers = [
        "tag_number", "name", "breed", "sex", "date_of_birth", "status",
        "weight_kg", "lactation_number", "dam_tag", "sire_tag",
        "purchase_date", "purchase_price", "notes",
    ]
    ws.append(headers)
    ws.append([
        "COW-001", "Rani", "Sahiwal", "female", "2022-03-15", "open",
        "380", "1", "", "", "2022-03-15", "150000", "Example row — delete before importing",
    ])
    return wb
