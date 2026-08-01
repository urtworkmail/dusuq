"""
Reports engine — generates Excel and PDF exports for all modules.
All report endpoints accept date_from / date_to query params.
"""
import io
from datetime import date, timedelta
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── Helpers ──────────────────────────────────────────────────────────────────

GREEN = "1A6B3C"
LIGHT_GREEN = "E8F5E9"
WHITE = "FFFFFF"
GREY = "F5F5F5"


def _style_header_row(ws, row_num, num_cols):
    fill = PatternFill("solid", fgColor=GREEN)
    font = Font(bold=True, color=WHITE)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


def _excel_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _write_title(ws, title, tenant_name):
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{tenant_name} — {title}"
    ws["A1"].font = Font(bold=True, size=14, color=GREEN)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated: {date.today()}"
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")


# ─── Reproduction Reports ─────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_inseminations(request):
    from apps.reproduction.models import Insemination
    tenant = request.tenant
    qs = Insemination.objects.filter(tenant=tenant).select_related("animal", "technician")
    date_from = request.query_params.get("date_from")
    date_to = request.query_params.get("date_to")
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    fmt = request.query_params.get("format", "json")
    rows = [
        {
            "date": str(r.date),
            "animal_tag": r.animal.tag_number,
            "animal_name": r.animal.name,
            "type": r.get_insemination_type_display(),
            "semen_batch": r.semen_batch,
            "bull_tag": r.bull_tag,
            "technician": r.technician.get_full_name() if r.technician else "",
            "vet_practitioner_number": r.veterinary_practitioner_number,
            "source_company": r.semen_source_company,
            "supplier_company": r.semen_supplier_company,
            "bull_breed": r.bull_breed,
            "donor_dam_breed": r.donor_dam_breed,
            "repeat": r.repeat_number,
            "expected_calving": str(r.expected_calving_date),
        }
        for r in qs
    ]

    if fmt == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inseminations"
        _write_title(ws, "Insemination Report", tenant.name)
        headers = [
            "Date", "Tag", "Name", "Type", "Semen/Embryo Batch", "Bull Tag", "Technician",
            "Vet Practitioner #", "Source Company", "Supplier Company", "Sire Breed",
            "Donor Dam Breed", "Repeat #", "Expected Calving",
        ]
        ws.append(headers)
        _style_header_row(ws, 3, len(headers))
        for r in rows:
            ws.append(list(r.values()))
        _auto_width(ws)
        return _excel_response(wb, f"inseminations_{date.today()}.xlsx")

    return Response({"count": len(rows), "results": rows})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_calvings(request):
    from apps.reproduction.models import Calving
    tenant = request.tenant
    qs = Calving.objects.filter(tenant=tenant).select_related("dam")
    date_from = request.query_params.get("date_from")
    date_to = request.query_params.get("date_to")
    if date_from:
        qs = qs.filter(calving_date__gte=date_from)
    if date_to:
        qs = qs.filter(calving_date__lte=date_to)

    fmt = request.query_params.get("format", "json")
    rows = [
        {
            "date": str(r.calving_date),
            "dam_tag": r.dam.tag_number,
            "dam_name": r.dam.name,
            "type": r.get_calving_type_display(),
            "dam_condition": r.dam_condition,
            "calf_tag": r.calf_tag,
            "calf_sex": r.calf_sex,
            "calf_weight_kg": str(r.calf_weight_kg or ""),
        }
        for r in qs
    ]

    if fmt == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Calvings"
        _write_title(ws, "Calving Report", tenant.name)
        headers = ["Date", "Dam Tag", "Dam Name", "Type", "Dam Condition", "Calf Tag", "Calf Sex", "Calf Weight (kg)"]
        ws.append(headers)
        _style_header_row(ws, 3, len(headers))
        for r in rows:
            ws.append(list(r.values()))
        _auto_width(ws)
        return _excel_response(wb, f"calvings_{date.today()}.xlsx")

    return Response({"count": len(rows), "results": rows})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_abortions(request):
    from apps.reproduction.models import Abortion
    tenant = request.tenant
    qs = Abortion.objects.filter(tenant=tenant).select_related("animal")
    date_from = request.query_params.get("date_from")
    date_to = request.query_params.get("date_to")
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    rows = [
        {
            "date": str(r.date),
            "animal_tag": r.animal.tag_number,
            "animal_name": r.animal.name,
            "gestation_days": r.gestation_stage_days or "",
            "cause": r.cause,
        }
        for r in qs
    ]

    fmt = request.query_params.get("format", "json")
    if fmt == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Abortions"
        _write_title(ws, "Abortion Report", tenant.name)
        headers = ["Date", "Tag", "Name", "Gestation Days", "Cause"]
        ws.append(headers)
        _style_header_row(ws, 3, len(headers))
        for r in rows:
            ws.append(list(r.values()))
        _auto_width(ws)
        return _excel_response(wb, f"abortions_{date.today()}.xlsx")

    return Response({"count": len(rows), "results": rows})


# ─── Health Reports ───────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_treatments(request):
    from apps.health.models import Treatment
    tenant = request.tenant
    qs = Treatment.objects.filter(tenant=tenant).select_related("animal", "administered_by")
    date_from = request.query_params.get("date_from")
    date_to = request.query_params.get("date_to")
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    rows = [
        {
            "date": str(r.date),
            "animal_tag": r.animal.tag_number,
            "diagnosis": r.diagnosis,
            "drug": r.drug,
            "dosage": r.dosage,
            "route": r.route,
            "withdrawal_days": r.withdrawal_days,
            "cost": str(r.cost),
            "outcome": r.outcome,
            "vet": r.administered_by.get_full_name() if r.administered_by else "",
        }
        for r in qs
    ]

    fmt = request.query_params.get("format", "json")
    if fmt == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Treatments"
        _write_title(ws, "Treatment Report", tenant.name)
        headers = ["Date", "Tag", "Diagnosis", "Drug", "Dosage", "Route", "Withdrawal Days", "Cost", "Outcome", "Vet"]
        ws.append(headers)
        _style_header_row(ws, 3, len(headers))
        for r in rows:
            ws.append(list(r.values()))
        _auto_width(ws)
        return _excel_response(wb, f"treatments_{date.today()}.xlsx")

    return Response({"count": len(rows), "results": rows})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_vaccinations(request):
    from apps.health.models import Vaccination
    tenant = request.tenant
    qs = Vaccination.objects.filter(tenant=tenant).select_related("animal", "shed")
    date_from = request.query_params.get("date_from")
    date_to = request.query_params.get("date_to")
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    rows = [
        {
            "date": str(r.date),
            "vaccine": r.vaccine_name,
            "batch": r.batch_number,
            "animal_tag": r.animal.tag_number if r.animal else "",
            "shed": r.shed.name if r.shed else "",
            "group": "Yes" if r.is_group_vaccination else "No",
            "next_due": str(r.next_due_date) if r.next_due_date else "",
            "cost": str(r.cost),
        }
        for r in qs
    ]

    fmt = request.query_params.get("format", "json")
    if fmt == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vaccinations"
        _write_title(ws, "Vaccination Report", tenant.name)
        headers = ["Date", "Vaccine", "Batch", "Animal Tag", "Shed", "Group?", "Next Due", "Cost"]
        ws.append(headers)
        _style_header_row(ws, 3, len(headers))
        for r in rows:
            ws.append(list(r.values()))
        _auto_width(ws)
        return _excel_response(wb, f"vaccinations_{date.today()}.xlsx")

    return Response({"count": len(rows), "results": rows})


# ─── Milk Reports ─────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_milk_daywise(request):
    from apps.milk.models import MilkRecord
    from django.db.models import Sum
    tenant = request.tenant
    date_from = request.query_params.get("date_from", str(date.today().replace(day=1)))
    date_to = request.query_params.get("date_to", str(date.today()))

    rows = list(
        MilkRecord.objects.filter(tenant=tenant, date__gte=date_from, date__lte=date_to)
        .values("date")
        .annotate(
            am=Sum("litres", filter=__import__("django.db.models", fromlist=["Q"]).Q(session="am")),
            pm=Sum("litres", filter=__import__("django.db.models", fromlist=["Q"]).Q(session="pm")),
            total=Sum("litres"),
        )
        .order_by("date")
    )

    fmt = request.query_params.get("format", "json")
    if fmt == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Milk"
        _write_title(ws, "Day-wise Milk Report", tenant.name)
        headers = ["Date", "AM (L)", "PM (L)", "Total (L)"]
        ws.append(headers)
        _style_header_row(ws, 3, len(headers))
        for r in rows:
            ws.append([str(r["date"]), float(r["am"] or 0), float(r["pm"] or 0), float(r["total"] or 0)])
        _auto_width(ws)
        return _excel_response(wb, f"milk_daywise_{date.today()}.xlsx")

    serializable = [
        {"date": str(r["date"]), "am": float(r["am"] or 0), "pm": float(r["pm"] or 0), "total": float(r["total"] or 0)}
        for r in rows
    ]
    return Response({"results": serializable})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_milk_animal(request):
    """Per-animal milk report for a date range."""
    from apps.milk.models import MilkRecord
    from django.db.models import Sum, Avg, Count
    tenant = request.tenant
    date_from = request.query_params.get("date_from", str(date.today().replace(day=1)))
    date_to = request.query_params.get("date_to", str(date.today()))

    rows = list(
        MilkRecord.objects.filter(tenant=tenant, date__gte=date_from, date__lte=date_to)
        .values("animal__tag_number", "animal__name")
        .annotate(
            total_litres=Sum("litres"),
            avg_daily=Avg("litres"),
            days_recorded=Count("date", distinct=True),
        )
        .order_by("-total_litres")
    )

    fmt = request.query_params.get("format", "json")
    if fmt == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Animal Milk"
        _write_title(ws, "Per-Animal Milk Report", tenant.name)
        headers = ["Tag", "Name", "Total (L)", "Avg Daily (L)", "Days Recorded"]
        ws.append(headers)
        _style_header_row(ws, 3, len(headers))
        for r in rows:
            ws.append([r["animal__tag_number"], r["animal__name"] or "", float(r["total_litres"]), round(float(r["avg_daily"]), 2), r["days_recorded"]])
        _auto_width(ws)
        return _excel_response(wb, f"milk_per_animal_{date.today()}.xlsx")

    return Response({"results": [
        {
            "tag": r["animal__tag_number"],
            "name": r["animal__name"] or "",
            "total_litres": float(r["total_litres"]),
            "avg_daily": round(float(r["avg_daily"]), 2),
            "days_recorded": r["days_recorded"],
        }
        for r in rows
    ]})


# ─── Finance Reports ──────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_transactions(request):
    from apps.accounts.models import Transaction
    tenant = request.tenant
    qs = Transaction.objects.filter(tenant=tenant).select_related("debit_account", "credit_account", "entered_by")
    date_from = request.query_params.get("date_from")
    date_to = request.query_params.get("date_to")
    txn_type = request.query_params.get("type")
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if txn_type:
        qs = qs.filter(transaction_type=txn_type)

    rows = [
        {
            "date": str(t.date),
            "type": t.get_transaction_type_display(),
            "reference": t.reference,
            "description": t.description,
            "debit_account": t.debit_account.name,
            "credit_account": t.credit_account.name,
            "amount": float(t.amount),
            "entered_by": t.entered_by.get_full_name() if t.entered_by else "",
        }
        for t in qs
    ]

    fmt = request.query_params.get("format", "json")
    if fmt == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        _write_title(ws, "Transaction Report", tenant.name)
        headers = ["Date", "Type", "Reference", "Description", "Debit Account", "Credit Account", "Amount", "Entered By"]
        ws.append(headers)
        _style_header_row(ws, 3, len(headers))
        for r in rows:
            ws.append(list(r.values()))
        _auto_width(ws)
        return _excel_response(wb, f"transactions_{date.today()}.xlsx")

    return Response({"count": len(rows), "results": rows})


# ─── Inventory Reports ────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_stock_summary(request):
    from apps.inventory.models import Product
    tenant = request.tenant
    category = request.query_params.get("category")
    products = Product.objects.filter(tenant=tenant, is_active=True)
    if category:
        products = products.filter(category=category)

    rows = [
        {
            "name": p.name,
            "category": p.category,
            "unit": p.unit,
            "current_stock": round(p.current_stock, 2),
            "reorder_level": float(p.reorder_level),
            "status": "Low" if p.current_stock <= float(p.reorder_level) else "OK",
        }
        for p in products
    ]

    fmt = request.query_params.get("format", "json")
    if fmt == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Summary"
        _write_title(ws, "Stock Summary Report", tenant.name)
        headers = ["Product", "Category", "Unit", "Current Stock", "Reorder Level", "Status"]
        ws.append(headers)
        _style_header_row(ws, 3, len(headers))
        for r in rows:
            row_data = list(r.values())
            ws.append(row_data)
            if r["status"] == "Low":
                row_idx = ws.max_row
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="FFCCCC")
        _auto_width(ws)
        return _excel_response(wb, f"stock_summary_{date.today()}.xlsx")

    return Response({"count": len(rows), "results": rows})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_consumption(request):
    from apps.inventory.models import Consumption
    from django.db.models import Sum
    tenant = request.tenant
    date_from = request.query_params.get("date_from", str(date.today().replace(day=1)))
    date_to = request.query_params.get("date_to", str(date.today()))
    category = request.query_params.get("category")

    qs = Consumption.objects.filter(tenant=tenant, date__gte=date_from, date__lte=date_to)
    if category:
        qs = qs.filter(product__category=category)

    rows = list(
        qs.values("product__name", "product__unit", "product__category")
        .annotate(total_consumed=Sum("quantity"))
        .order_by("product__category", "-total_consumed")
    )

    fmt = request.query_params.get("format", "json")
    if fmt == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consumption"
        _write_title(ws, "Consumption Report", tenant.name)
        headers = ["Product", "Unit", "Category", "Total Consumed"]
        ws.append(headers)
        _style_header_row(ws, 3, len(headers))
        for r in rows:
            ws.append([r["product__name"], r["product__unit"], r["product__category"], float(r["total_consumed"])])
        _auto_width(ws)
        return _excel_response(wb, f"consumption_{date.today()}.xlsx")

    return Response({"results": [
        {
            "product": r["product__name"],
            "unit": r["product__unit"],
            "category": r["product__category"],
            "total_consumed": float(r["total_consumed"]),
        }
        for r in rows
    ]})


# ─── Forecasting (deterministic — calculated from real data, not AI) ─────────

def _linear_trend(points):
    """
    Least-squares slope/intercept for [(x, y), ...] using pure Python (no
    numpy dependency). Returns (slope, intercept); (0, mean_y) if fewer than
    2 distinct x values, since a trend line is meaningless with less data.
    """
    n = len(points)
    if n < 2:
        return 0.0, (points[0][1] if points else 0.0)
    sum_x = sum(p[0] for p in points)
    sum_y = sum(p[1] for p in points)
    mean_x = sum_x / n
    mean_y = sum_y / n
    num = sum((p[0] - mean_x) * (p[1] - mean_y) for p in points)
    den = sum((p[0] - mean_x) ** 2 for p in points)
    if den == 0:
        return 0.0, mean_y
    slope = num / den
    intercept = mean_y - slope * mean_x
    return slope, intercept


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def forecast_summary(request):
    """
    Real, calculated forecasts — not AI-generated text. Counts actual
    upcoming due dates from reproduction records, and projects milk trend via
    linear regression over the trailing 90 days of daily totals.
    """
    from django.db.models import Sum
    from apps.reproduction.models import Insemination
    from apps.milk.models import MilkRecord

    tenant = request.tenant
    today = date.today()
    horizon_days = int(request.query_params.get("horizon_days", 30))
    horizon = today + timedelta(days=horizon_days)

    pregnant_inseminations = list(
        Insemination.objects.filter(tenant=tenant, animal__status="pregnant").select_related("animal")
    )

    expected_calvings = []
    expected_dry_offs = []
    for ins in pregnant_inseminations:
        ecd = ins.expected_calving_date
        if today <= ecd <= horizon:
            expected_calvings.append({"animal_tag": ins.animal.tag_number, "expected_date": str(ecd)})
        dry_off_date = ecd - timedelta(days=60)
        if today <= dry_off_date <= horizon:
            expected_dry_offs.append({"animal_tag": ins.animal.tag_number, "expected_date": str(dry_off_date)})

    # Milk trend — daily totals over the trailing 90 days, linear regression
    # projected forward to the end of the horizon.
    trend_start = today - timedelta(days=90)
    daily_totals = list(
        MilkRecord.objects.filter(tenant=tenant, date__gte=trend_start, date__lte=today)
        .values("date").annotate(total=Sum("litres")).order_by("date")
    )
    points = [((row["date"] - trend_start).days, float(row["total"])) for row in daily_totals]
    slope, intercept = _linear_trend(points)

    current_daily_avg = round(sum(p[1] for p in points) / len(points), 1) if points else 0.0
    projected_x = (horizon - trend_start).days
    projected_daily = round(slope * projected_x + intercept, 1) if points else 0.0

    return Response({
        "horizon_days": horizon_days,
        "expected_calvings": {
            "count": len(expected_calvings),
            "animals": expected_calvings,
        },
        "expected_dry_offs": {
            "count": len(expected_dry_offs),
            "animals": expected_dry_offs,
        },
        "milk_trend": {
            "current_daily_avg_litres": current_daily_avg,
            "projected_daily_avg_litres": max(projected_daily, 0.0),
            "trend_direction": "up" if slope > 0.05 else ("down" if slope < -0.05 else "flat"),
            "based_on_days": len(points),
        },
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def herd_growth_projection(request):
    """
    Simple, transparent herd-size projection: current headcount plus the
    trailing-12-month net growth rate (calvings minus deaths/sold/culled),
    extrapolated linearly to +12/+24/+36 months. Deliberately not a
    compounding/demographic model — easy to explain and sanity-check, which
    matters more here than modeling precision for a first version.
    """
    from apps.animals.models import Animal, AnimalStatus
    from apps.reproduction.models import Calving

    tenant = request.tenant
    today = date.today()
    year_ago = today - timedelta(days=365)

    current_herd_size = Animal.objects.filter(tenant=tenant, is_active=True).count()

    calvings_last_year = Calving.objects.filter(
        tenant=tenant, calving_date__gte=year_ago, calving_date__lte=today
    ).exclude(calf_sex="stillborn").count()

    losses_last_year = Animal.objects.filter(
        tenant=tenant,
        status__in=[AnimalStatus.SOLD, AnimalStatus.DEAD, AnimalStatus.CULLED],
        updated_at__date__gte=year_ago, updated_at__date__lte=today,
    ).count()

    avg_monthly_calvings = calvings_last_year / 12
    avg_monthly_losses = losses_last_year / 12
    net_monthly_growth = avg_monthly_calvings - avg_monthly_losses

    projections = {
        f"plus_{months}_months": max(round(current_herd_size + net_monthly_growth * months), 0)
        for months in (12, 24, 36)
    }

    return Response({
        "current_herd_size": current_herd_size,
        "avg_monthly_calvings": round(avg_monthly_calvings, 2),
        "avg_monthly_losses": round(avg_monthly_losses, 2),
        "net_monthly_growth": round(net_monthly_growth, 2),
        "projections": projections,
        "note": (
            "Linear projection based on the trailing 12 months' calving and "
            "loss (sold/dead/culled) rate. Does not account for seasonality, "
            "purchases, or changes in herd management."
        ),
    })
