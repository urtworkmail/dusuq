import io
from datetime import date, timedelta
from decimal import Decimal

from django.db.models import Sum, Avg, Count
from django.http import HttpResponse
from rest_framework import generics, filters, status
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend
import openpyxl

from apps.animals.models import Animal
from .models import MilkRecord, ConsumptionHead, MilkConsumption, MilkDispatch
from .serializers import (
    MilkRecordSerializer, BulkMilkRecordSerializer,
    ConsumptionHeadSerializer, MilkConsumptionSerializer, MilkDispatchSerializer,
)


class MilkRecordListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = MilkRecordSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["animal", "date", "session"]
    ordering = ["-date", "animal__tag_number"]

    def get_queryset(self):
        qs = MilkRecord.objects.filter(tenant=self.request.tenant).select_related("animal")
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        return qs

    def perform_create(self, serializer):
        serializer.save(tenant=self.request.tenant, recorded_by=self.request.user)


class MilkRecordDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = MilkRecordSerializer

    def get_queryset(self):
        return MilkRecord.objects.filter(tenant=self.request.tenant)


class BulkMilkEntryView(APIView):
    """POST a list of records — creates or updates by (animal, date, session)."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = BulkMilkRecordSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        objects = serializer.create(serializer.validated_data)
        return Response(
            {"saved": len(objects), "detail": f"{len(objects)} records saved."},
            status=status.HTTP_200_OK,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def milk_entry_sheet(request):
    """Returns list of milking animals ready for bulk entry on a given date+session."""
    today_param = request.query_params.get("date", str(date.today()))
    session = request.query_params.get("session", "am")

    milking_animals = Animal.objects.filter(
        tenant=request.tenant,
        is_active=True,
        status__in=["open", "inseminated", "pregnant"],
    ).order_by("tag_number")

    existing = {
        r.animal_id: r
        for r in MilkRecord.objects.filter(
            tenant=request.tenant, date=today_param, session=session
        )
    }

    rows = []
    for animal in milking_animals:
        rec = existing.get(animal.id)
        rows.append({
            "animal_id": animal.id,
            "tag_number": animal.tag_number,
            "name": animal.name,
            "shed": animal.shed.name if animal.shed else "",
            "record_id": rec.id if rec else None,
            "litres": float(rec.litres) if rec else None,
            "fat_percent": float(rec.fat_percent) if rec and rec.fat_percent else None,
            "snf_percent": float(rec.snf_percent) if rec and rec.snf_percent else None,
        })

    return Response({"date": today_param, "session": session, "animals": rows})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def milk_excel_template(request):
    """Download a pre-filled Excel template for milk entry."""
    session = request.query_params.get("session", "am")
    entry_date = request.query_params.get("date", str(date.today()))

    animals = Animal.objects.filter(
        tenant=request.tenant, is_active=True,
        status__in=["open", "inseminated", "pregnant"],
    ).order_by("tag_number")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Milk Entry"
    ws.append(["animal_id", "tag_number", "name", "date", "session", "litres", "fat_percent", "snf_percent"])
    for a in animals:
        ws.append([a.id, a.tag_number, a.name or "", entry_date, session, "", "", ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="milk_entry_{entry_date}_{session}.xlsx"'
    return resp


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser])
def milk_excel_import(request):
    """Import milk records from uploaded Excel file."""
    file = request.FILES.get("file")
    if not file:
        return Response({"detail": "No file provided."}, status=400)

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
    except Exception as e:
        return Response({"detail": f"Invalid Excel file: {e}"}, status=400)

    headers = [cell.value for cell in ws[1]]
    required = {"animal_id", "date", "session", "litres"}
    if not required.issubset(set(h for h in headers if h)):
        return Response({"detail": f"Missing columns. Required: {required}"}, status=400)

    saved, errors = 0, []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = dict(zip(headers, row))
        if not row_data.get("animal_id") or not row_data.get("litres"):
            continue
        try:
            animal = Animal.objects.get(id=row_data["animal_id"], tenant=request.tenant)
            MilkRecord.objects.update_or_create(
                tenant=request.tenant,
                animal=animal,
                date=row_data["date"],
                session=str(row_data["session"]).lower(),
                defaults={
                    "litres": Decimal(str(row_data["litres"])),
                    "fat_percent": row_data.get("fat_percent") or None,
                    "snf_percent": row_data.get("snf_percent") or None,
                    "recorded_by": request.user,
                },
            )
            saved += 1
        except Animal.DoesNotExist:
            errors.append(f"Row {i}: Animal ID {row_data['animal_id']} not found.")
        except Exception as e:
            errors.append(f"Row {i}: {e}")

    return Response({"saved": saved, "errors": errors})


# ─── Consumption Heads ────────────────────────────────────────────────────────

class ConsumptionHeadListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ConsumptionHeadSerializer

    def get_queryset(self):
        return ConsumptionHead.objects.filter(tenant=self.request.tenant)

    def perform_create(self, serializer):
        serializer.save(tenant=self.request.tenant)


class ConsumptionHeadDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ConsumptionHeadSerializer

    def get_queryset(self):
        return ConsumptionHead.objects.filter(tenant=self.request.tenant)


# ─── Milk Consumption ─────────────────────────────────────────────────────────

class MilkConsumptionListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = MilkConsumptionSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["head", "date"]
    ordering = ["-date"]

    def get_queryset(self):
        qs = MilkConsumption.objects.filter(tenant=self.request.tenant).select_related("head")
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        return qs

    def perform_create(self, serializer):
        serializer.save(tenant=self.request.tenant)


class MilkConsumptionDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = MilkConsumptionSerializer

    def get_queryset(self):
        return MilkConsumption.objects.filter(tenant=self.request.tenant)


# ─── Milk Dispatch ────────────────────────────────────────────────────────────

class MilkDispatchListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = MilkDispatchSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["dispatch_type", "date"]
    search_fields = ["buyer_name"]
    ordering = ["-date"]

    def get_queryset(self):
        qs = MilkDispatch.objects.filter(tenant=self.request.tenant)
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        return qs

    def perform_create(self, serializer):
        serializer.save(tenant=self.request.tenant)


class MilkDispatchDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = MilkDispatchSerializer

    def get_queryset(self):
        return MilkDispatch.objects.filter(tenant=self.request.tenant)


# ─── Dashboard ────────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def milk_dashboard(request):
    tenant = request.tenant
    today = date.today()
    month_start = today.replace(day=1)

    milking_count = Animal.objects.filter(
        tenant=tenant, is_active=True, status__in=["open", "inseminated", "pregnant"]
    ).count()

    # Today's total
    today_total = MilkRecord.objects.filter(
        tenant=tenant, date=today
    ).aggregate(total=Sum("litres"))["total"] or 0

    # This month totals per day
    daily_totals = list(
        MilkRecord.objects.filter(tenant=tenant, date__gte=month_start)
        .values("date")
        .annotate(total=Sum("litres"))
        .order_by("date")
    )

    # Monthly total
    month_total = sum(d["total"] for d in daily_totals)

    # Per-animal avg this month
    per_animal = list(
        MilkRecord.objects.filter(tenant=tenant, date__gte=month_start)
        .values("animal__tag_number", "animal__name")
        .annotate(total=Sum("litres"), avg_daily=Avg("litres"))
        .order_by("-total")[:20]
    )

    # Chiller balance today
    produced_today = float(today_total)
    consumed_today = float(
        MilkConsumption.objects.filter(tenant=tenant, date=today)
        .aggregate(total=Sum("litres"))["total"] or 0
    )
    dispatched_today = float(
        MilkDispatch.objects.filter(tenant=tenant, date=today)
        .aggregate(total=Sum("gross_litres"))["total"] or 0
    )
    chiller_balance = produced_today - consumed_today - dispatched_today

    return Response({
        "milking_animals": milking_count,
        "today_total_litres": float(today_total),
        "month_total_litres": float(month_total),
        "chiller_balance_today": chiller_balance,
        "daily_totals": [{"date": str(d["date"]), "total": float(d["total"])} for d in daily_totals],
        "top_animals": per_animal,
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def chiller_status(request):
    """Running chiller balance for a date range."""
    tenant = request.tenant
    date_from = request.query_params.get("date_from", str(date.today().replace(day=1)))
    date_to = request.query_params.get("date_to", str(date.today()))

    rows = []
    current = date.fromisoformat(date_from)
    end = date.fromisoformat(date_to)
    while current <= end:
        produced = float(
            MilkRecord.objects.filter(tenant=tenant, date=current)
            .aggregate(t=Sum("litres"))["t"] or 0
        )
        consumed = float(
            MilkConsumption.objects.filter(tenant=tenant, date=current)
            .aggregate(t=Sum("litres"))["t"] or 0
        )
        dispatched = float(
            MilkDispatch.objects.filter(tenant=tenant, date=current)
            .aggregate(t=Sum("gross_litres"))["t"] or 0
        )
        rows.append({
            "date": str(current),
            "produced": produced,
            "consumed": consumed,
            "dispatched": dispatched,
            "balance": produced - consumed - dispatched,
        })
        current += timedelta(days=1)

    return Response(rows)
